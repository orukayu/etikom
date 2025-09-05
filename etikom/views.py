from django.shortcuts import render
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.urls import reverse

from .models import Stok
from .models import Siparis
from .models import Blog
from .models import Kargo
from .models import Iade
from .models import Gider
from .forms import StokFormu
from .forms import KayitFormu
from .forms import GirisFormu
from .forms import SiparisFormu
from .forms import KargoFormu
from .forms import IadeFormu
from .forms import GiderFormu

from django.db.models import Sum
from django.db.models import Q
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout

import pandas as pd         # pip install pandas ile excel yukleme icin pandas kuruldu ve pd kisaltma adi verildi, ayrıca pip install openpyxl ında kurulması gerekıyor.
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
import os
from io import BytesIO
from django.conf import settings

from django.db.models import Min, Max

from django.db.models.functions import ExtractWeek
from django.db.models.functions import ExtractMonth
from django.db.models.functions import ExtractYear
from datetime import datetime, timedelta
import calendar
import json

from django.contrib import messages

# Create your views here.

def girisyap(request):
    title = 'Etikom'
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('raporlarurl')  # Ana sayfa URL'inizi buraya yazın
        else:
            messages.error(request, 'Kullanıcı adı veya şifre yanlış.')
    context = {
        'title': title
    }

    return render(request, 'etikom/giris.html', context)

def stokharaketleriyap(request, sort=None):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    tfta = Stok.objects.filter(Firmaadi=firma_adi_id).values('Afaturano').order_by('Afaturano').distinct().count()   # toplam fatura sayisi
    tsc = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').order_by('Stokkodu').distinct().count()      # Toplam stok cesidi sayısı

    if tsc == 0:
        tstg = 0
        ksa = 0
        tstm = 0
        ostm = 0
        tsts = 0
        tstc = 0
    else:
        tstg = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0).aggregate(Sum('Adet'))["Adet__sum"]
        ksa = Stok.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Adet"))["Adet__sum"]
        tstm = Stok.objects.filter(Firmaadi=firma_adi_id, Tur='A', Toplam__gt=0).aggregate(Sum("Toplam"))["Toplam__sum"]

        # None kontrolü
        if tstg is None:
            tstg = 0
        if ksa is None:
            ksa = 0
        if tstm is None:
            tstm = 0

        tstc = abs(ksa - tstg)                                                                                      # Stok cikisi

        if tstc == 0:
            tsatis = 0
        else:
            tsatis = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__lte=0).aggregate(Sum("Toplam"))["Toplam__sum"]          # toplam stok satis bedeli

        tsts = abs(tsatis)

        ostm = (tstm - tsts) / ksa

    stsys = Stok.objects.filter(Firmaadi=firma_adi_id).count()

    if sort == 'az-tur':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Tur').values()
    elif sort == 'za-tur':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Tur').values()
    elif sort == 'az-fatura-no':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Afaturano').values()
    elif sort == 'za-fatura-no':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Afaturano').values()
    elif sort == 'az-tarih':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Alistarihi').values()
    elif sort == 'za-tarih':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Alistarihi').values()
    elif sort == 'az-stok-kodu':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Stokkodu').values()
    elif sort == 'za-stok-kodu':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Stokkodu').values()
    elif sort == 'az-adet':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Adet').values()
    elif sort == 'za-adet':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Adet').values()
    elif sort == 'az-fiyat':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Alisfiyati').values()
    elif sort == 'za-fiyat':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Alisfiyati').values()
    elif sort == 'az-toplam':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('Toplam').values()
    elif sort == 'za-toplam':
        stok = Stok.objects.filter(Firmaadi=firma_adi_id).order_by('-Toplam').values()
    else:
        stok = Stok.objects.filter(Firmaadi=firma_adi_id)

    firma = request.user.username
    title = 'Stok Hareketleri'

    context = {
        'stok': stok,
        'firma_adi': firma_adi,
        'title': title,
        'stsys': stsys,
        'ostm': ostm,
        'tstm': tstm,
        'tsts': tsts,
        'tfta': tfta,
        'tsc': tsc,
        'tstg': tstg,
        'tstc': tstc,
        'ksa': ksa,
        'firma': firma,
    }
    

    return render(request, 'etikom/stokharaketleri.html', context)


def siparislistesiyap(request, sort=None):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    tpys = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Pazaryeri').order_by('Pazaryeri').distinct().count()
    tsps = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Siparisno').order_by('Siparisno').distinct().count()   # toplam siparis sayisi
    tsts = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').order_by('Stokkodu').distinct().count()
    tstc = Siparis.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0).aggregate(Sum('Adet'))["Adet__sum"]                # tum siparislerdeki toplam urun adedi

    if tstc is None:
        tstc = 0

    tsius = Siparis.objects.filter(Firmaadi=firma_adi_id, Adet__lte=0).aggregate(Sum('Adet'))["Adet__sum"]                # tum siparislerdeki iade urun sayısı

    if tsius is None:
        ius = 0
    else:
        ius = abs(tsius)

    if tstc == 0:                   
        orsp = 0
    else:
        orsp = tstc / tsps          # siparis basina dusen urun sayisi

    tstt = Siparis.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Toplam"))["Toplam__sum"]

    if tstt is None:
        tstt = 0
        ostt = 0
    else:
        ostt = tstt / tsps

    stsys = Siparis.objects.filter(Firmaadi=firma_adi_id).count()

    if sort == 'az-tur':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Tur').values()
    elif sort == 'za-tur':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Tur').values()
    elif sort == 'az-pazaryeri':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Pazaryeri').values()
    elif sort == 'za-pazaryeri':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Pazaryeri').values()
    elif sort == 'az-tarih':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Tarih').values()
    elif sort == 'za-tarih':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Tarih').values()
    elif sort == 'az-siparis-no':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Siparisno').values()
    elif sort == 'za-siparis-no':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Siparisno').values()
    elif sort == 'az-stok-kodu':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Stokkodu').values()
    elif sort == 'za-stok-kodu':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Stokkodu').values()
    elif sort == 'az-adet':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Adet').values()
    elif sort == 'za-adet':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Adet').values()
    elif sort == 'az-satis-fiyati':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Satisfiyati').values()
    elif sort == 'za-satis-fiyati':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Satisfiyati').values()
    elif sort == 'az-toplam':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Toplam').values()
    elif sort == 'za-toplam':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Toplam').values()
    elif sort == 'az-komisyon-orani':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Komisyon').values()
    elif sort == 'za-komisyon-orani':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Komisyon').values()
    elif sort == 'az-komisyon-tutari':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Komisyontutari').values()
    elif sort == 'za-komisyon-tutari':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Komisyontutari').values()
    else:
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id)

    firma = request.user.username
    title = 'Sipariş Listesi'

    context = {
        'siparis': siparis,
        'firma_adi': firma_adi,
        'title': title,
        'tpys': tpys,
        'tsps': tsps,
        'tsts': tsts,
        'tstc': tstc,
        'ius': ius,
        'orsp': orsp,
        'tstt': tstt,
        'ostt': ostt,
        'stsys': stsys,
        'firma': firma,
    }

    return render(request, 'etikom/siparislistesi.html', context)

def kargolistesiyap(request, sort=None):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).count()
    if kargo is None:
        kargo_sayisi = 0
    else:
        kargo_sayisi = kargo

    desiler = Kargo.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Desi"))["Desi__sum"]                       # toplam desi
    tutarlar = Kargo.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Kargotutari"))["Kargotutari__sum"]        # toplam kargo
    hizmetler = Kargo.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Hizmetbedeli"))["Hizmetbedeli__sum"]     # toplam hizmet
    islemler = Kargo.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Islembedeli"))["Islembedeli__sum"]     # toplam islem
    toplamlar = Kargo.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Toplam"))["Toplam__sum"]                 # genel toplam


    if kargo_sayisi == 0:
        ort_desi = 0
        top_tutar = 0
        ort_tutar = 0
        ort_hizmet = 0
        top_hizmet = 0
        ort_islem = 0
        top_islem = 0
        gen_toplam = 0
    else:
        ort_desi = desiler / kargo
        top_tutar = tutarlar
        ort_tutar = top_tutar / kargo_sayisi
        top_hizmet = hizmetler
        ort_hizmet = top_hizmet / kargo_sayisi
        top_islem = islemler
        ort_islem = top_islem / kargo_sayisi
        gen_toplam = toplamlar
    

    if sort == 'az-tur':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Tur').values()
    elif sort == 'za-tur':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Tur').values()
    elif sort == 'az-siparis-no':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Siparisno').values()
    elif sort == 'za-siparis-no':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Siparisno').values()
    elif sort == 'az-desi':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Desi').values()
    elif sort == 'za-desi':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Desi').values()
    elif sort == 'az-kargo-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Kargotutari').values()
    elif sort == 'za-kargo-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Kargotutari').values()
    elif sort == 'az-hizmet-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Hizmetbedeli').values()
    elif sort == 'za-hizmet-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Hizmetbedeli').values()
    elif sort == 'az-islem-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Islembedeli').values()
    elif sort == 'za-islem-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Islembedeli').values()
    elif sort == 'az-toplam-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('Toplam').values()
    elif sort == 'za-toplam-tutari':
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id).order_by('-Toplam').values()
    else:
        kargo = Kargo.objects.filter(Firmaadi=firma_adi_id)

    firma = request.user.username
    title = 'Kargo Listesi'

    context = {
        'kargo': kargo,
        'firma_adi': firma_adi,
        'firma': firma,
        'title': title,
        'kargo_sayisi': kargo_sayisi,
        'desiler': desiler,
        'ort_desi': ort_desi,
        'top_tutar': top_tutar,
        'ort_tutar': ort_tutar,
        'top_hizmet': top_hizmet,
        'ort_hizmet': ort_hizmet,
        'top_islem': top_islem,
        'ort_islem': ort_islem,
        'gen_toplam': gen_toplam,
    }
    

    return render(request, 'etikom/kargolistesi.html', context)

def raporlaryap(request):
    if not request.user.is_authenticated:  # Eğer kullanıcı giriş yapmamışsa
        demo_kullanici = authenticate(request, username='demo', password='demodemo')
        if demo_kullanici is not None:
            login(request, demo_kullanici)
        else:
            return redirect('girisurl')

    firma_adi = request.user.username
    firma_adi_id = request.user.id

    title = 'Genel Raporlar'   

    kont_tststt = Stok.objects.filter(Firmaadi=firma_adi_id, Tur='T', Adet__lte=0).aggregate(Sum("Toplam"))["Toplam__sum"]       # toptan satis tutari
    if kont_tststt == None:
        tststt = 0
    else:
        tststt = abs(kont_tststt)

    kont_tsistt = Siparis.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Toplam"))["Toplam__sum"]     # toplam siparis tutari
    if kont_tsistt == None:
        tsistt = 0
    else:
        tsistt = abs(kont_tsistt)

    ts = tststt + tsistt

    # Donem sayisi icin en küçük ve en büyük tarihleri bulma
    tarih_araligi = Siparis.objects.filter(Firmaadi=firma_adi_id).aggregate(en_kucuk_tarih=Min('Tarih'), en_buyuk_tarih=Max('Tarih'))

    if tsistt == 0:
        en_kucuk_tarih = datetime.today()
        en_buyuk_tarih = datetime.today()
    else:
        en_kucuk_tarih = tarih_araligi['en_kucuk_tarih']
        en_buyuk_tarih = tarih_araligi['en_buyuk_tarih']

    # Tarihler arasındaki farkı hesaplama
    ekt = en_kucuk_tarih
    ekt_ay = ekt.month
    ekt_yil = ekt.year

    ebt = en_buyuk_tarih
    ebt_ay = ebt.month
    ebt_yil = ebt.year

    if ebt_yil - ekt_yil == 0:
        tdns = ebt.month - ekt.month + 1
    elif ebt_yil - ekt_yil == 1:
        tdns = (12 - ekt_ay) + ebt_ay
    else:
        tdns = (12 - ekt_ay) + ebt_ay + (12 * (ebt_yil - ekt_yil - 1))
    
    ort_dnm_sts = ts / tdns  # Aylik ortalama satis

    n_s_t = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__lt=0).values('Stokkodu').annotate(total_cikis=Sum('Adet')).order_by('total_cikis')[:5] #en çok satılan 5 ürün
    n_s_t = list(n_s_t)  # QuerySet'i listeye çevir

    n_s_t_list = list(n_s_t)
    for item in n_s_t_list:
        item['total_cikis'] = abs(item['total_cikis'])

    # Eksik sayıda item varsa, boş item ekle
    n_s_t_count = len(n_s_t)
    if n_s_t_count < 5:
        for _ in range(5 - n_s_t_count):
            n_s_t.append({'Stokkodu': None, 'total_adet': 0})

    n_s_y = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__lt=0).values('Stokkodu').annotate(total_cikis=Sum('Adet')).order_by('-total_cikis')[:5] #en az satilan 5 urun
    n_s_y = list(n_s_y)  # QuerySet'i listeye çevir

    n_s_y_list = list(n_s_y)
    for item in n_s_y_list:
        item['total_cikis'] = abs(item['total_cikis'])


    # Eksik sayıda item varsa, boş item ekle
    n_s_y_count = len(n_s_y)
    if n_s_y_count < 5:
        for _ in range(5 - n_s_y_count):
            n_s_y.append({'Stokkodu': None, 'total_adet': 0})

    s_t_u = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').annotate(total_adet=Sum('Adet')).filter(total_adet__lte=0)[:5] #stok 0 olan urunler
    s_t_u = list(s_t_u)  # QuerySet'i listeye çevir

    # Eksik sayıda item varsa, boş item ekle
    s_t_u_count = len(s_t_u)
    if s_t_u_count < 5:
        for _ in range(5 - s_t_u_count):
            s_t_u.append({'Stokkodu': None, 'total_adet': 0})

    s_a_u = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').annotate(total_adet=Sum('Adet')).filter(total_adet__gt=0).order_by('total_adet')[:5] #stogu azalan urunler
    c_p_y = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Pazaryeri').order_by('Pazaryeri').distinct()  # pazaryerleri listesi
    b_p_y = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Pazaryeri').annotate(total_satis=Sum('Toplam')).order_by('-total_satis')[:5] #en buyuk pazaryeri
    s_y_u = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').annotate(total_adet=Sum('Adet')).filter(total_adet__gt=0).order_by('-total_adet')[:5] #stogu azalan urunler

    today = datetime.today()    # Şu anki tarih
    buay = today.month
    buyil = today.year
    bugun = today.weekday()   # Pazartesi=0, Pazar=6 olacak şekilde bugünün rakamı

    # Bu haftanın siparişlerini filtreleme
    haftabasi = today - timedelta(days=today.weekday())  # Haftanın başlangıcı (Pazartesi)
    haftasonu = haftabasi + timedelta(days=6)  # Haftanın bitişi (Pazar)
    buhaftaki_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih__range=[haftabasi, haftasonu])
    bh_siplerin_toplami = buhaftaki_sipler.aggregate(bhsip_tt=Sum('Toplam'))['bhsip_tt']
    if bh_siplerin_toplami is None:
        bh_siplerin_toplami = 0

    # Gecen haftanın siparişlerini filtreleme
    gecen_haftasonu = today - timedelta(days=bugun+1)
    gecen_haftabasi = gecen_haftasonu - timedelta(days=6)
    gcn_haftaki_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih__range=[gecen_haftabasi, gecen_haftasonu])
    gh_siplerin_toplami = gcn_haftaki_sipler.aggregate(ghsip_tt=Sum('Toplam'))['ghsip_tt']
    if gh_siplerin_toplami is None:
        gh_siplerin_toplami = 0

    # Bu ayın siparişlerini filtreleme
    aybasi = datetime(buyil, buay, 1)
    ayinsongunu = calendar.monthrange(buyil, buay)[1]
    aysonu = datetime(buyil, buay, ayinsongunu)
    buayki_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih__range=[aybasi, aysonu])
    ba_sipler_toplami = buayki_sipler.aggregate(basip_tt=Sum('Toplam'))['basip_tt']
    if ba_sipler_toplami is None:
        ba_sipler_toplami = 0

    # Gecen ayın siparişlerini filtreleme
    gcn_aybasi = datetime(buyil, buay-1, 1)
    gcn_ayinsongunu = calendar.monthrange(buyil, buay-1)[1]
    gcn_aysonu = datetime(buyil, buay-1, gcn_ayinsongunu)
    gcn_ayki_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih__range=[gcn_aybasi, gcn_aysonu])
    ga_sipler_toplami = gcn_ayki_sipler.aggregate(gasip_tt=Sum('Toplam'))['gasip_tt']
    if ga_sipler_toplami is None:
        ga_sipler_toplami = 0

    # Bu yılın siparişlerini filtreleme
    yilbasi = datetime(buyil, 1, 1)
    yilsonu = datetime(buyil, 12, 31)
    buyilki_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih__range=[yilbasi, yilsonu])
    by_sipler_toplami = buyilki_sipler.aggregate(bysip_tt=Sum('Toplam'))['bysip_tt']
    if by_sipler_toplami is None:
        by_sipler_toplami = 0

    # Gecen yılın siparişlerini filtreleme
    gcn_yilbasi = datetime(buyil-1, 1, 1)
    gcn_yilsonu = datetime(buyil-1, 12, 31)
    gcn_yilki_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih__range=[gcn_yilbasi, gcn_yilsonu])
    gy_sipler_toplami = gcn_yilki_sipler.aggregate(gysip_tt=Sum('Toplam'))['gysip_tt']
    if gy_sipler_toplami is None:
        gy_sipler_toplami = 0


    tsc = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').order_by('Stokkodu').distinct().count()      # Toplam stok cesidi sayısı

    if tsc == 0:
        tstg = 0
        ksa = 0
        tstm = 0
        ostm = 0
    else:
        tstg = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0).aggregate(Sum('Adet'))["Adet__sum"]
        ksa = Stok.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Adet"))["Adet__sum"]
        tstm = Stok.objects.filter(Firmaadi=firma_adi_id, Toplam__gt=0).aggregate(Sum("Toplam"))["Toplam__sum"]
        ostm = tstm / tstg

        # None kontrolü
        if tstg is None:
            tstg = 0
            ostm = 0
        if ksa is None:
            ksa = 0
            ostm = 0
        if tstm is None:
            tstm = 0
            ostm = 0

    tstc = abs(ksa - tstg)

    edkocek = Siparis.objects.filter(Firmaadi=firma_adi_id, Tur='S').values('Komisyon').order_by('Komisyon')
    if edkocek:
        edko = edkocek[0]['Komisyon']
    else:
        edko = 0

    eykocek = Siparis.objects.filter(Firmaadi=firma_adi_id, Tur='S').values('Komisyon').order_by('-Komisyon')
    if eykocek:
        eyko = eykocek[0]['Komisyon']
    else:
        eyko = 0

    kom_topla = Siparis.objects.filter(Firmaadi=firma_adi_id, Tur='S').aggregate(Sum("Komisyon"))["Komisyon__sum"]
    sip_say = Siparis.objects.filter(Firmaadi=firma_adi_id, Tur='S').count()

    if sip_say == 0:
        edko = 0
        eyko = 0
        oko = 0
        tokotu = 0
    else:
        oko = kom_topla / sip_say

    tokotu = Siparis.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Komisyontutari"))["Komisyontutari__sum"]
    if tokotu is None:
        tokotu = 0

    n_i_u = Iade.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').annotate(total_cikis=Sum('Adet')).order_by('-total_cikis')[:5] #en çok iade 5 ürün
    n_i_u = list(n_i_u)  # QuerySet'i listeye çevir

    n_i_u_list = list(n_i_u)
    for item in n_i_u_list:
        item['total_cikis'] = (item['total_cikis'])

    # Eksik sayıda item varsa, boş item ekle
    n_i_u_count = len(n_i_u)
    if n_i_u_count < 5:
        for _ in range(5 - n_i_u_count):
            n_i_u.append({'Stokkodu': None, 'total_cikis': 0})


    # 1. Iade tablosundaki benzersiz stok kodlarını alın
    iade_stok_kodlari = Iade.objects.filter(Firmaadi=firma_adi_id).values_list('Stokkodu', flat=True).distinct()

    # 2. Iadesi olmayan stok kodlarını Siparis tablosunda bulun
    iadesi_olmayan_stok_kodlari = Siparis.objects.filter(Firmaadi=firma_adi_id).exclude(Stokkodu__in=iade_stok_kodlari)

    # 3. Adetleri toplayın ve sıralayın
    adet_toplamlari = (
        iadesi_olmayan_stok_kodlari
        .values('Stokkodu')
        .annotate(total_adet=Sum('Adet'))
        .order_by('-total_adet')
    )

    # 4. İlk 5 öğeyi seçin
    ilk_5_oge = adet_toplamlari[:5]

    ilk_5 = list(ilk_5_oge)  # QuerySet'i listeye çevir

    ilk_5_list = list(ilk_5)
    for item in ilk_5_list:
        item['total_adet'] = (item['total_adet'])

    # Eksik sayıda item varsa, boş item ekle
    ilk_5_count = len(ilk_5)
    if ilk_5_count < 5:
        for _ in range(5 - ilk_5_count):
            ilk_5.append({'Stokkodu': None, 'total_adet': 0})

    giderler = Gider.objects.filter(Firmaadi=firma_adi_id).values('Baslik').annotate(total_gider=Sum('Tutar')).order_by('-total_gider')[:5]
    giderler = list(giderler)  # QuerySet'i listeye çevir

    giderler_list = list(giderler)
    for item in giderler_list:
        item['total_gider'] = (item['total_gider'])

    # Eksik sayıda item varsa, boş item ekle
    giderler_count = len(giderler)
    if giderler_count < 5:
        for _ in range(5 - giderler_count):
            giderler.append({'Baslik': None, 'total_gider': 0})


    krg_ve_islemler = Kargo.objects.filter(Firmaadi=firma_adi_id, Tur='K').aggregate(Sum("Toplam"))["Toplam__sum"]

    edkrcek = Kargo.objects.filter(Firmaadi=firma_adi_id, Tur='K').values('Kargotutari').order_by('Kargotutari')
    if edkrcek:
        edkt = edkrcek[0]['Kargotutari']
    else:
        edkt = 0

    eykrcek = Kargo.objects.filter(Firmaadi=firma_adi_id, Tur='K').values('Kargotutari').order_by('-Kargotutari')
    if eykrcek:
        eykt = eykrcek[0]['Kargotutari']
    else:
        eykt = 0

    krg_topla = Kargo.objects.filter(Firmaadi=firma_adi_id, Tur='K').aggregate(Sum("Kargotutari"))["Kargotutari__sum"]
    if krg_topla is None:
        krg_topla = 0
        krg_ve_islemler = 0

    krg_say = Kargo.objects.filter(Firmaadi=firma_adi_id, Tur='K').count()

    if krg_say == 0:
        edkt = 0
        eykt = 0
        okt = 0
        tokrtu = 0
    else:
        okt = krg_topla / krg_say

    
    giderler_kontrol = Gider.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Tutar"))["Tutar__sum"]
    if giderler_kontrol is None:
        giderler_toplami = 0
    else:
        giderler_toplami = giderler_kontrol


    iadeler_kontrol = Iade.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Iadetutari"))["Iadetutari__sum"]
    if iadeler_kontrol is None:
        iadeler_toplami = 0
    else:
        iadeler_toplami = iadeler_kontrol

    tgider = tstm + krg_ve_islemler + giderler_toplami + iadeler_toplami + tokotu

    kalan = ts - tgider

    context = {
        'title': title,
        'firma_adi': firma_adi,
        'tdns': tdns,
        'tsistt': tsistt,
        'tststt': tststt,
        'ts': ts,
        'tstm': tstm,
        'ort_dnm_sts': ort_dnm_sts,
        'n_s_t': n_s_t,
        'n_s_y': n_s_y,
        's_t_u': s_t_u,
        's_a_u': s_a_u,
        'c_p_y': c_p_y,
        'b_p_y': b_p_y,
        's_y_u': s_y_u,
        'today': today,
        'bh_siplerin_toplami': bh_siplerin_toplami,
        'gh_siplerin_toplami': gh_siplerin_toplami,
        'ba_sipler_toplami': ba_sipler_toplami,
        'ga_sipler_toplami': ga_sipler_toplami,
        'by_sipler_toplami': by_sipler_toplami,
        'gy_sipler_toplami': gy_sipler_toplami,
        'tsc': tsc,
        'tstg': tstg,
        'tstc': tstc,
        'ksa': ksa,
        'edko': edko,
        'eyko': eyko,
        'oko': oko,
        'tokotu': tokotu,
        'n_i_u': n_i_u,
        'ilk_5': ilk_5,
        'giderler': giderler,
        'edkt': edkt,
        'eykt': eykt,
        'okt': okt,
        'krg_topla': krg_topla,
        'tgider': tgider,
        'kalan': kalan,
    }

    return render(request, 'etikom/raporlar.html', context)

def kayitol(request):
    if not request.user.is_authenticated:  # Eğer kullanıcı giriş yapmamışsa
        demo_kullanici = authenticate(request, username='demo', password='demodemo')
        if demo_kullanici is not None:
            login(request, demo_kullanici)
        else:
            return redirect('girisurl')

    title = 'Kayıt Ol'
    kayit = KayitFormu()
    firma_adi = request.user.username
    if request.method == "POST":
        kayit = KayitFormu(request.POST)

        firma_adi = request.POST["firma_adi"]
        email = request.POST["email"]
        password1 = request.POST["password1"]
        password2 = request.POST["password2"]

        if kayit.is_valid():
            if password1 == password2:
                var = User.objects.filter(username=firma_adi).count()
                if var > 0:
                    kayit.add_error('firma_adi', 'Bu firma adı zaten kullanılıyor.')
                else:
                    kayit = User.objects.create_user(username=firma_adi, email=email, password=password1)
                    kayit.save()
                    return redirect('cikisurl')
            else:
                kayit.add_error('password1', 'Şifre farklı girilmiş.')

    return render(request, 'etikom/kayitol.html', {'firma_adi': firma_adi, 'kayit': kayit, 'title': title})



def cikisyap(request):
    logout(request)
    return redirect('girisurl')


def blogyap(request):
    if not request.user.is_authenticated:  # Eğer kullanıcı giriş yapmamışsa
        demo_kullanici = authenticate(request, username='demo', password='demodemo')
        if demo_kullanici is not None:
            login(request, demo_kullanici)
        else:
            return redirect('girisurl')

    firma_adi = request.user.username
    title = 'Blog'
    blog = Blog.objects.all()
    # ... iletişim sayfası içeriğini oluşturun
    return render(request, 'etikom/blog.html', {'title': title, 'firma_adi': firma_adi, 'blog': blog})

def blogdetayyap(request, url):
    firma_adi = request.user.username
    title = 'Blog Detayı'
    blog = Blog.objects.filter(Url=url)
    # ... iletişim sayfası içeriğini oluşturun
    return render(request, 'etikom/blogdetay.html', {'title': title, 'firma_adi': firma_adi, 'blog': blog})

def iletisimyap(request):
    if not request.user.is_authenticated:  # Eğer kullanıcı giriş yapmamışsa
        demo_kullanici = authenticate(request, username='demo', password='demodemo')
        if demo_kullanici is not None:
            login(request, demo_kullanici)
        else:
            return redirect('girisurl')

    firma_adi = request.user.username
    title = 'İletişim'
    # ... iletişim sayfası içeriğini oluşturun
    return render(request, 'etikom/iletisim.html', {'title': title, 'firma_adi': firma_adi})

def hakkimizdayap(request):
    if not request.user.is_authenticated:  # Eğer kullanıcı giriş yapmamışsa
        demo_kullanici = authenticate(request, username='demo', password='demodemo')
        if demo_kullanici is not None:
            login(request, demo_kullanici)
        else:
            return redirect('girisurl')

    firma_adi = request.user.username
    title = 'Hakkımızda'
    # ... iletişim sayfası içeriğini oluşturun
    return render(request, 'etikom/hakkimizda.html', {'title': title, 'firma_adi': firma_adi})

def fiyatlamayap(request):
    firma_adi = request.user.username
    title = 'Fiyatlandırma'
    # ... iletişim sayfası içeriğini oluşturun
    return render(request, 'etikom/fiyatlandirma.html', {'title': title, 'firma_adi': firma_adi})

def stokexceliyuklemeyap(request):
    title = 'Excel Yükle'
    firma_adi = request.user.username

    if request.method == "POST":
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                tur = 'A' if row['Adet'] > 0 else 'T'
                stok = Stok(
                    Afaturano = row['Fatura No'],
                    Alistarihi = row['Fatura Tarihi'],
                    Stokkodu = row['Stok Kodu'],
                    Adet = row['Adet'],
                    Alisfiyati = row['Fiyat'],
                    Toplam = row['Toplam'],
                    Tur = tur,
                    Firmaadi = request.user
                )
                stok.save1()
            
            return redirect('stoklistesiurl')

    return render(request, 'etikom/stokexceliyukle.html', {'firma_adi': firma_adi, 'title': title})


def sipexceliyuklemeyap(request):
    title = 'Excel Yükle'
    firma_adi = request.user.username

    if request.method == "POST":
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                sip = Siparis(
                    Pazaryeri = row['Pazaryeri'],
                    Tarih = row['Sipariş Tarihi'],
                    Siparisno = row['Sipariş No'],
                    Stokkodu = row['Stok Kodu'],
                    Adet = row['Adet'],
                    Satisfiyati = row['Satış Fiyatı'],
                    Komisyon = row['Komisyon (%)'],
                    Firmaadi = request.user,
                    Tur = 'S'
                )
                sip.save3()
                stk = Stok(
                    Firmaadi = request.user,
                    Afaturano = row['Sipariş No'],
                    Stokkodu = row['Stok Kodu'],
                    Adet = row['Adet'] * -1,
                    Alisfiyati = row['Satış Fiyatı'],
                    Tur = 'S'
                )
                stk.save1()

            return redirect('siparislistesiurl')
    
    return render(request, 'etikom/sipexceliyukle.html', {'firma_adi': firma_adi, 'title': title})

def kargoexceliyuklemeyap(request):
    title = 'Excel Yükle'
    firma_adi = request.user.username

    if request.method == "POST":
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                siparis_no = row['Sipariş No']
                if isinstance(siparis_no, int):
                    siparis = str(siparis_no)
                else:
                    siparis = siparis_no
                kargo = Kargo(
                    Desi = row['Desi'],
                    Kargotutari = row['Kargo Tutarı'],
                    Hizmetbedeli = row['Hizmet Bedeli'],
                    Islembedeli = row['İşlem Bedeli'],
                    Siparisno = siparis,
                    Firmaadi = request.user,
                    Tur = 'K'
                )
                kargo.save5()
            
            return redirect('kargolistesiurl')

    return render(request, 'etikom/kargoexceliyukle.html', {'firma_adi': firma_adi, 'title': title})


def stokexceliindir(request):
    firma_adi_id = request.user.id

    # Stok modelinden tüm verileri al
    stoklar = Stok.objects.filter(Firmaadi=firma_adi_id)

    # Stok verilerini bir DataFrame'e dönüştür
    data = {
        'Belge No': [stok.Afaturano for stok in stoklar],
        'Stok Kodu': [stok.Stokkodu for stok in stoklar],
        'Adet': [stok.Adet for stok in stoklar],
        'Fiyat': [str(stok.Alisfiyati).replace('.', ',') for stok in stoklar],
        'Toplam': [str(stok.Toplam).replace('.', ',') for stok in stoklar],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)  # Buffer'ın başına git

    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_stok_hareketleri.xlsx"'

    return response



def stokharaketdetayiyap(request, firma, pk):
    
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if firma_adi != firma:
        return redirect('cikisurl')

    stokturcek = Stok.objects.filter(id=pk).values_list('Tur', flat=True).first()

    if stokturcek == 'A':
        kontrolstok = get_object_or_404(Stok, pk=pk)

        if request.method == "POST":
            form = StokFormu(request.POST, instance=kontrolstok)
            if 'stoksil' in request.POST:
                kontrolstok.delete()
                return redirect('stoklistesiurl')
            elif 'stokekle' in request.POST:
                if form.is_valid():
                    post = form.save(commit=False)
                    post.Firmaadi = request.user
                    if post.Adet > 0:
                        post.Tur = 'A'
                    else:
                        post.Tur = 'T'
                    post.save1()
                    return redirect('stoklistesiurl')
        else:
            form = StokFormu(instance=kontrolstok)

    elif stokturcek == 'T':
        kontrolstok = get_object_or_404(Stok, pk=pk)

        if request.method == "POST":
            form = StokFormu(request.POST, instance=kontrolstok)
            if 'stoksil' in request.POST:
                kontrolstok.delete()
                return redirect('stoklistesiurl')
            elif 'stokekle' in request.POST:
                if form.is_valid():
                    post = form.save(commit=False)
                    post.Firmaadi = request.user
                    if post.Adet > 0:
                        post.Tur = 'A'
                    else:
                        post.Tur = 'T'
                    post.save1()
                    return redirect('stoklistesiurl')
        else:
            form = StokFormu(instance=kontrolstok)

    elif stokturcek == 'S':
        bncek = Stok.objects.filter(id=pk).values_list('Afaturano', flat=True).first()
        skcek = Stok.objects.filter(id=pk).values_list('Stokkodu', flat=True).first()
        pk = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=bncek, Stokkodu=skcek, Tur='S').values_list('id', flat=True).first()

        return redirect('siparisduzeltmeurl', firma, pk)

    elif stokturcek == 'İ':
        bncek = Stok.objects.filter(id=pk).values_list('Afaturano', flat=True).first()
        skcek = Stok.objects.filter(id=pk).values_list('Stokkodu', flat=True).first()
        pk = Iade.objects.filter(Firmaadi=firma_adi_id, Siparisno=bncek, Stokkodu=skcek, Tur='İ').values_list('id', flat=True).first()

        return redirect('iadeduzeltmeurl', firma, pk)

    else:
        form = StokFormu(instance=kontrolstok)


    title = 'Stok Detayı'
    form = StokFormu(instance=kontrolstok)
    return render(request, 'etikom/stokharaketdetayi.html', {'form': form, 'firma_adi': firma_adi, 'title': title})



def stoklistesiyap(request):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    tscs = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').order_by('Stokkodu').distinct().count()
    toplamalimlar = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0, Tur='A').aggregate(toplam_adet=Sum('Adet'))['toplam_adet']
    toplamkalanlar = Stok.objects.filter(Firmaadi=firma_adi_id).aggregate(toplam_adet=Sum('Adet'))['toplam_adet']
    toplamtoplamlar = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0, Tur='A').aggregate(toplam_adet=Sum('Toplam'))['toplam_adet']
    toplamadetler = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0, Tur='A').aggregate(toplam_adet=Sum('Adet'))['toplam_adet']

    if toplamtoplamlar is None:
        oaf = 0
    else:
        oaf = toplamtoplamlar / toplamadetler    

    if tscs == 0:
        tsc = 0
        toplamalimlar = 0
        toplamkalanlar = 0
    else:
        tsc = tscs


    # Adet > 0 olan satırlardan Toplam sütunlarının toplamını hesaplama
    toplam_toplam = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0, Tur='A').values('Stokkodu').annotate(total_toplam=Sum('Toplam')).order_by('Stokkodu')

    # Adet > 0 olan satırlardan Adet sütunlarının toplamını hesaplama
    toplam_adet_filtered = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0, Tur='A').values('Stokkodu').annotate(total_adet_filtered=Sum('Adet')).order_by('Stokkodu')

    # Stok Kodu'na göre tüm Adet sütunlarının toplamını hesaplama (Adet değerine bakmaksızın)
    toplam_adet_all = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').annotate(total_adet_all=Sum('Adet')).order_by('Stokkodu')

    # Sonuçları birleştirerek tek bir yapı oluşturma
    etopla = {}
    for item in toplam_toplam:
        etopla[item['Stokkodu']] = {
            'total_toplam': item['total_toplam'],
            'total_adet_filtered': 0,
            'total_adet_all': 0,
            'ortalama_alisfiyati': 0
        }

    for item in toplam_adet_filtered:
        if item['Stokkodu'] in etopla:
            etopla[item['Stokkodu']]['total_adet_filtered'] = item['total_adet_filtered']

    for item in toplam_adet_all:
        if item['Stokkodu'] in etopla:
            etopla[item['Stokkodu']]['total_adet_all'] = item['total_adet_all']

    # Ortalama alış fiyatını hesaplama
    for stok_kodu, values in etopla.items():
        total_toplam = values['total_toplam']
        total_adet_filtered = values['total_adet_filtered']
        values['ortalama_alisfiyati'] = total_toplam / total_adet_filtered if total_adet_filtered else 0

    title = 'Stok Listesi'

    context = {
        'etopla': etopla,
        'firma_adi': firma_adi,
        'title': title,
        'tsc': tsc,
        'tsga': toplamalimlar,
        'gsa': toplamkalanlar,
        'oaf': oaf,
    }

    return render(request, 'etikom/stoklistesi.html', context)


def sayimexcelindir(request):
    firma_adi_id = request.user.id

    # Stok modelinden tüm verileri al
    stoklar = Stok.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').annotate(total_adet=Sum('Adet'))
    toplam_giris = Stok.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0).values('Stokkodu').annotate(toplam_giris=Sum('Adet')).order_by('Stokkodu')

    # Stok verilerini bir DataFrame'e dönüştür
    data = {
        'Stok Kodu': [item['Stokkodu'] for item in stoklar],
        'Toplam Giriş': [item['toplam_giris'] for item in toplam_giris],
        'Mevcut Adet': [item['total_adet'] for item in stoklar],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)  # Buffer'ın başına git

    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_stok_adetleri.xlsx"'

    return response


def stokeklemeyap(request):

    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if request.method == "POST":
        if 'stokekle' in request.POST:
            form = StokFormu(request.POST)
            if form.is_valid():
                post = form.save(commit=False)
                post.Firmaadi = request.user
                if post.Adet > 0:
                    post.Tur = 'A'
                else:
                    post.Tur = 'T'
                post.save1()
                return redirect('stoklistesiurl')

    else:
        form = StokFormu()

    title = 'Stok Ekle'
    
    return render(request, 'etikom/stokekle.html', {'form': form, 'firma_adi': firma_adi, 'title': title})

def fiyathesaplamayap(request):

    firma_adi = request.user.username
    firma_adi_id = request.user.id

    title = 'Fiyat Hesaplama'
    
    return render(request, 'etikom/fiyathesaplama.html', {'firma_adi': firma_adi, 'title': title})


def sipariseklemeyap(request):

    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if request.method == "POST":
        if 'sipekle' in request.POST:
            siparis = SiparisFormu(request.POST, user=request.user)
            if siparis.is_valid():
                post = siparis.save(commit=False)
                post.Firmaadi = request.user
                post.Tur = 'S'
                post.save3()

                sipno = siparis.cleaned_data['Siparisno']
                stokkodu = siparis.cleaned_data['Stokkodu']
                sayi = siparis.cleaned_data['Adet']
                adet = sayi * -1
                satfiyat = siparis.cleaned_data['Satisfiyati']
                Firmaadi = request.user

                # Book kaydet
                stok = Stok(Firmaadi=Firmaadi, Afaturano=sipno, Stokkodu=stokkodu, Adet=adet, Alisfiyati=satfiyat, Tur='S')
                stok.save1()
                return redirect('siparislistesiurl')

    else:
        siparis = SiparisFormu(user=request.user)

    title = 'Sipariş Ekle'
    
    return render(request, 'etikom/siparisekleme.html', {'siparis': siparis, 'firma_adi': firma_adi, 'title': title})


def stokfaturasiyap(request, sort):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Fatura Detayı'

    faturaturcek = Stok.objects.filter(Firmaadi=firma_adi_id, Afaturano=sort).values_list('Tur', flat=True).first()

    if faturaturcek == 'S':
        return redirect('siparisdetayurl', sort)
    elif faturaturcek == 'İ':
        return redirect('siparisdetayurl', sort)


    fatura = Stok.objects.filter(Firmaadi=firma_adi_id, Afaturano=sort)

    tstc = Stok.objects.filter(Firmaadi=firma_adi_id, Afaturano=sort).values('Stokkodu').order_by('Stokkodu').distinct().count()
    tsta = Stok.objects.filter(Firmaadi=firma_adi_id, Afaturano=sort).aggregate(Sum("Adet"))["Adet__sum"]
    
    tsta = abs(tsta)

    sftt = Stok.objects.filter(Firmaadi=firma_adi_id, Afaturano=sort).aggregate(Sum("Toplam"))["Toplam__sum"]

    sftt = abs(sftt)

    oafi = sftt / tsta

    context = {
        'firma_adi': firma_adi,
        'title': title,
        'sort': sort,
        'fatura': fatura,
        'tstc': tstc,
        'tsta': tsta,
        'sftt': sftt,
        'oafi': oafi,
    }

    return render(request, 'etikom/stokfaturasi.html', context)



def stokgecmisiyap(request, sort):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Stok Geçmişi'

    stok = Stok.objects.filter(Firmaadi=firma_adi_id, Stokkodu=sort)
    
    ma = Stok.objects.filter(Firmaadi=firma_adi_id, Stokkodu=sort).aggregate(Sum("Adet"))["Adet__sum"]                      # mevcut adet
    tstc = Stok.objects.filter(Firmaadi=firma_adi_id, Stokkodu=sort, Adet__lte=0).aggregate(Sum("Adet"))["Adet__sum"]       # toplam stok cikisi
    tstg = Stok.objects.filter(Firmaadi=firma_adi_id, Stokkodu=sort, Adet__gt=0).aggregate(Sum("Adet"))["Adet__sum"]        # toplam stok girisi
    
    if tstc == None:
        tstc = 0
    else:
        tstc = abs(tstc)

    tsaf = Stok.objects.filter(Firmaadi=firma_adi_id, Stokkodu=sort, Adet__gt=0).aggregate(Sum("Toplam"))["Toplam__sum"]    # toplam stok alis toplami

    if tstg == None:
        soaf = 0
    else:
        soaf = tsaf / (ma + tstc)

    tssf = Stok.objects.filter(Firmaadi=firma_adi_id, Stokkodu=sort, Adet__lte=0).aggregate(Sum("Toplam"))["Toplam__sum"]   # toplam stok satis toplami
    if tssf == None:
        tssf = 0

    if tstc != 0:
        sosf = tssf / tstc
        sosf = abs(sosf)
    else:
        sosf = 0

    context = {
        'firma_adi': firma_adi,
        'title': title,
        'sort': sort,
        'stok': stok,
        'ma': ma,
        'tstc': tstc,
        'soaf': soaf,
        'sosf': sosf,
    }

    return render(request, 'etikom/stokgecmisi.html', context)

def siparisduzeltme(request, firma, pk):

    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if firma_adi != firma:
        return redirect('cikisurl')

    sipturcek = Siparis.objects.filter(id=pk).values_list('Tur', flat=True).first()

    if sipturcek == 'S':
        kontrolsiparis = get_object_or_404(Siparis, pk=pk)
        siparis = SiparisFormu(request.POST, instance=kontrolsiparis, user=request.user)
        if 'siparissil' in request.POST:
            stok_entry = Stok.objects.filter(Afaturano=kontrolsiparis.Siparisno, Stokkodu=kontrolsiparis.Stokkodu, Tur='S').first()
            if stok_entry:
                stok_entry.delete()

            kontrolsiparis.delete()
            return redirect('siparislistesiurl')
        elif 'siparisekle' in request.POST:
            stok_entry = Stok.objects.filter(Afaturano=kontrolsiparis.Siparisno, Stokkodu=kontrolsiparis.Stokkodu, Tur='S').first()
            if stok_entry:
                stok_entry.delete()

            kontrolsiparis.delete()
            if siparis.is_valid():
                post = siparis.save(commit=False)
                post.Firmaadi = request.user
                post.Tur = 'S'
                post.save3()


                Firmaadi = request.user
                adet = siparis.cleaned_data['Adet']
                fyt = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=kontrolsiparis.Siparisno, Stokkodu=kontrolsiparis.Stokkodu, Tur='S').values('Satisfiyati').first()
                fiyat = fyt['Satisfiyati']
                stok = Stok(Firmaadi=request.user, Afaturano=kontrolsiparis.Siparisno, Stokkodu=kontrolsiparis.Stokkodu, Adet=adet, Alisfiyati=fiyat, Tur='S')
                stok.save1()

                return redirect('siparislistesiurl')
        else:
            siparis = SiparisFormu(instance=kontrolsiparis, user=request.user)

    elif sipturcek == 'İ':
        sipcek = Siparis.objects.filter(id=pk).values_list('Siparisno', flat=True).first()
        skcek = Siparis.objects.filter(id=pk).values_list('Stokkodu', flat=True).first()
        pk = Iade.objects.filter(Firmaadi=firma_adi_id, Siparisno=sipcek, Stokkodu=skcek, Tur='İ').values_list('id', flat=True).first()

        return redirect('iadeduzeltmeurl', firma, pk)

    else:
        siparis = SiparisFormu(instance=kontrolsiparis, user=request.user)

    title = 'Sipariş Detayı'
    siparis = SiparisFormu(instance=kontrolsiparis, user=request.user)
    return render(request, 'etikom/siparisduzeltme.html', {'siparis': siparis, 'firma_adi': firma_adi, 'title': title})


def sipexceliindir(request):
    firma_adi_id = request.user.id

    # Siparis modelinden tüm verileri al
    sipler = Siparis.objects.filter(Firmaadi=firma_adi_id)

    # Siparis verilerini bir DataFrame'e dönüştür
    data = {
        'Pazaryeri': [siparis.Pazaryeri for siparis in sipler],
        'Tarih': [siparis.Tarih.strftime('%d.%m.%Y') for siparis in sipler],
        'Sipariş No': [siparis.Siparisno for siparis in sipler],
        'Stok Kodu': [siparis.Stokkodu for siparis in sipler],
        'Adet': [str(siparis.Adet).replace('.', ',') for siparis in sipler],
        'Satış Fiyatı': [str(siparis.Satisfiyati).replace('.', ',') for siparis in sipler],
        'Toplam': [str(siparis.Toplam).replace('.', ',') for siparis in sipler],
        'Komisyon %': [str(siparis.Komisyon).replace('.', ',') for siparis in sipler],
        'Komisyon Tutarı': [str(siparis.Komisyontutari).replace('.', ',') for siparis in sipler],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Siparisler')
    excel_buffer.seek(0)  # Buffer'ın başına git

    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_siparisler.xlsx"'

    return response


def pazaryeridetayiyap(request, pzr):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Pazaryeri Detayı'

    pazaryeri = Siparis.objects.filter(Firmaadi=firma_adi_id, Pazaryeri=pzr)

    tstc = Siparis.objects.filter(Firmaadi=firma_adi_id, Pazaryeri=pzr).values('Stokkodu').order_by('Stokkodu').distinct().count()
    tsta = Siparis.objects.filter(Firmaadi=firma_adi_id, Pazaryeri=pzr).aggregate(Sum("Adet"))["Adet__sum"]
    
    tsta = abs(tsta)

    sftt = Siparis.objects.filter(Firmaadi=firma_adi_id, Pazaryeri=pzr).aggregate(Sum("Toplam"))["Toplam__sum"]

    sftt = abs(sftt)

    oafi = sftt / tsta

    context = {
        'firma_adi': firma_adi,
        'title': title,
        'pazaryeri': pazaryeri,
        'tstc': tstc,
        'tsta': tsta,
        'sftt': sftt,
        'oafi': oafi,
        'pzr': pzr,
    }

    return render(request, 'etikom/pazaryeridetay.html', context)

def siparisleritoplayap(request, sira):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Sipariş Topla'

    tarihler = list(Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Tarih').values_list('Tarih', flat=True).distinct())

    liste_sayisi = len(tarihler)

    if liste_sayisi == 0:
        return redirect('siparisyokurl')
    else:
        en_son_tarih = tarihler[sira-1]

    if en_son_tarih is not None:
        # En son tarihe ait siparişleri filtreleyin
        en_son_sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih=en_son_tarih)
        
        # Stok kodlarını ve adet toplamlarını gruplayın ve hesaplayın
        stok_toplamlari = en_son_sipler.values('Stokkodu').annotate(total_quantity=Sum('Adet'))
    else:
        stok_toplamlari = []

    context = {
        'firma_adi': firma_adi,
        'title': title,
        'en_son_tarih': en_son_tarih,
        'stok_toplamlari': stok_toplamlari,
        'sira': sira,
        'liste_sayisi': liste_sayisi,
    }

    return render(request, 'etikom/siparisleritopla.html', context)

def siparisyok(request):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Sipariş Yok'

    context = {
        'firma_adi': firma_adi,
        'title': title,
    }

    return render(request, 'etikom/siparisyok.html', context)

def siparistoplaexceli(request, sira):
    firma_adi_id = request.user.id

    # Tarihleri sıralı ve benzersiz hale getirin
    tarihler = list(Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Tarih').values_list('Tarih', flat=True).distinct())
    en_son_tarih = tarihler[sira-1]

    # Tarihi d.m.y formatında biçimlendirin
    en_son_tarih_str = en_son_tarih.strftime("%d.%m.%Y")

    sipler = Siparis.objects.filter(Firmaadi=firma_adi_id, Tarih=en_son_tarih)
    stok_toplamlari = sipler.values('Stokkodu').annotate(total_quantity=Sum('Adet'))

    # Siparis verilerini bir DataFrame'e dönüştür
    data = {
        'Stok Kodu': [stok['Stokkodu'] for stok in stok_toplamlari],
        'Toplam Adet': [stok['total_quantity'] for stok in stok_toplamlari],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1)
    excel_buffer.seek(0)  # Buffer'ın başına git

    # openpyxl ile çalışma kitabını yükleyin
    workbook = load_workbook(excel_buffer)
    sheet = workbook.active

    # A1 hücresine en son tarihi yazın
    sheet.merge_cells('A1:B1')
    sheet['A1'] = f"{en_son_tarih_str} Tarihli Sipariş İçerikleri"

    # Dosyayı tekrar buffer'a kaydet
    response_buffer = BytesIO()
    workbook.save(response_buffer)
    response_buffer.seek(0)


    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(response_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_{en_son_tarih_str}_sevkiyat.xlsx"'

    return response

def kargoeklemeyap(request):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Kargo'
    
    if request.method == "POST":
        if 'kargoekle' in request.POST:
            form = KargoFormu(request.POST)
            if form.is_valid():
                sip_no = form.cleaned_data['Siparisno']
                kontrol = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no).count()
                if kontrol > 0:
                    post = form.save(commit=False)
                    post.Firmaadi = request.user
                    post.Tur = 'K'
                    post.save5()
                    return redirect('kargolistesiurl')
                else:
                    form.add_error('Siparisno', 'Sipariş No Mevcut Değil !')
                    
    else:
        form = KargoFormu()

    return render(request, 'etikom/kargoekle.html', {'title': title, 'firma_adi': firma_adi, 'form': form})

def kargoduzeltme(request, firma, pk):
    
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if firma_adi != firma:
        return redirect('cikisurl')

    turcek = Kargo.objects.filter(pk=pk).values_list('Tur', flat=True).first()

    if turcek == 'K':
        kontrolkargo = get_object_or_404(Kargo, pk=pk)

        if request.method == "POST":
            form = KargoFormu(request.POST, instance=kontrolkargo)
            if 'kargosil' in request.POST:
                kontrolkargo.delete()
                return redirect('kargolistesiurl')
            elif 'kargoekle' in request.POST:
                if form.is_valid():
                    post = form.save(commit=False)
                    post.Firmaadi = request.user
                    post.Tur = 'K'
                    post.save5()
                    return redirect('kargolistesiurl')
        else:
            form = KargoFormu(instance=kontrolkargo)

    elif turcek == 'İ':
        sipcek = Kargo.objects.filter(id=pk).values_list('Siparisno', flat=True).first()
        skcek = Kargo.objects.filter(id=pk).values_list('Stokkodu', flat=True).first()
        pk = Iade.objects.filter(Firmaadi=firma_adi_id, Siparisno=sipcek, Stokkodu=skcek, Tur='İ').values_list('id', flat=True).first()

        return redirect('iadeduzeltmeurl', firma, pk)

    else:
        form = KargoFormu(instance=kontrolkargo)

    title = 'Kargo Detayı'

    return render(request, 'etikom/kargoduzeltme.html', {'form': form, 'firma_adi': firma_adi, 'title': title})

def kargoexceliindir(request):
    firma_adi_id = request.user.id

    # Kargo modelinden tüm verileri al
    kargolar = Kargo.objects.filter(Firmaadi=firma_adi_id)

    # Kargo verilerini bir DataFrame'e dönüştür
    data = {
        'Sipariş No': [kargo.Siparisno for kargo in kargolar],
        'Desi': [kargo.Desi for kargo in kargolar],
        'Kargo Tutarı': [str(kargo.Kargotutari).replace('.', ',') for kargo in kargolar],
        'Hizmet + İşlem Bedeli': [str(kargo.Hizmetbedeli).replace('.', ',') for kargo in kargolar],
        'Toplam': [str(kargo.Toplam).replace('.', ',') for kargo in kargolar],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)  # Buffer'ın başına git

    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_giden_kargo_listesi.xlsx"'

    return response


def iadeeklemeyap(request):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'İade Kargo'
    
    if request.method == "POST":
        if 'iadeekle' in request.POST:
            form = IadeFormu(request.POST)
            if form.is_valid():
                sip_no = form.cleaned_data['Siparisno']
                sip_kontrol = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no).count()
                if sip_kontrol > 0:
                    stk_no = form.cleaned_data['Stokkodu']
                    stk_kontrol = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no, Stokkodu=stk_no).count()
                    if stk_kontrol > 0:
                        post = form.save(commit=False)
                        post.Firmaadi = request.user
                        post.Tur = 'İ'
                        post.save6()

                        Firmaadi = request.user
                        adet = form.cleaned_data['Adet']
                        fyt = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no, Stokkodu=stk_no).values('Satisfiyati').first()
                        fiyat = fyt['Satisfiyati']
                        stok = Stok(Firmaadi=Firmaadi, Afaturano=sip_no, Stokkodu=stk_no, Adet=adet, Alisfiyati=fiyat, Tur='İ')
                        stok.save1()

                        pzr_yeri = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no, Stokkodu=stk_no).values('Pazaryeri').first()
                        pazaryeri = pzr_yeri['Pazaryeri']
                        adt = form.cleaned_data['Adet'] * -1
                        trh = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no, Stokkodu=stk_no).values('Tarih').first()
                        tarih = trh['Tarih']
                        kmsyn = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sip_no, Stokkodu=stk_no).values('Komisyon').first()
                        komisyon = kmsyn['Komisyon']
                        siparis = Siparis(Firmaadi=Firmaadi, Pazaryeri=pazaryeri, Tarih=tarih, Siparisno=sip_no, Stokkodu=stk_no, Adet=adt, Komisyon=komisyon, Satisfiyati=fiyat, Tur='İ')
                        siparis.save4()

                        desi = form.cleaned_data['Desi']
                        iade_ttr = form.cleaned_data['Iadetutari']
                        kargo = Kargo(Firmaadi=Firmaadi, Tur='İ', Siparisno=sip_no, Stokkodu=stk_no, Desi=desi, Hizmetbedeli=0, Islembedeli=0, Kargotutari=iade_ttr)
                        kargo.save5()

                        return redirect('iadelistesiurl')
                    else:
                        form.add_error('Stokkodu', 'Stok Kodu Siparişe Ait Değil !')
                else:
                    form.add_error('Siparisno', 'Sipariş No Mevcut Değil !')
                    
    else:
        form = IadeFormu()

    return render(request, 'etikom/iadeekle.html', {'title': title, 'firma_adi': firma_adi, 'form': form})

def iadelistesiyap(request, sort=None):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    iade = Iade.objects.filter(Firmaadi=firma_adi_id).count()
    if iade is None:
        iade_sayisi = 0
    else:
        iade_sayisi = iade

    desiler = Iade.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Desi"))["Desi__sum"]                            # toplam desi
    tutarlar = Iade.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Iadetutari"))["Iadetutari__sum"]               # toplam kargo
    stoklar = Iade.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').order_by('Stokkodu').distinct().count()     # stok çeşidi
    adetler = Iade.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Adet"))["Adet__sum"]                            # toplam adet


    if iade_sayisi == 0:
        ort_desi = 0
        top_tutar = 0
        ort_tutar = 0
        top_stok = 0
        top_hizmet = 0
        gen_toplam = 0
        top_adet = 0
    else:
        ort_desi = desiler / iade
        top_tutar = tutarlar
        ort_tutar = top_tutar / iade_sayisi
        top_stok = stoklar
        top_adet = adetler
    

    if sort == 'az-tur':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('Tur').values()
    elif sort == 'za-tur':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('-Tur').values()
    elif sort == 'az-siparis-no':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('Siparisno').values()
    elif sort == 'za-siparis-no':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('-Siparisno').values()
    elif sort == 'az-desi':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('Desi').values()
    elif sort == 'za-desi':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('-Desi').values()
    elif sort == 'az-iade-tutari':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('Iadetutari').values()
    elif sort == 'za-iade-tutari':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('-Iadetutari').values()
    elif sort == 'az-adet':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('Adet').values()
    elif sort == 'za-adet':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('-Adet').values()
    elif sort == 'az-stok-kodu':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('Stokkodu').values()
    elif sort == 'za-stok-kodu':
        iade = Iade.objects.filter(Firmaadi=firma_adi_id).order_by('-Stokkodu').values()
    else:
        iade = Iade.objects.filter(Firmaadi=firma_adi_id)

    firma = request.user.username
    title = 'İade Listesi'

    context = {
        'iade': iade,
        'firma_adi': firma_adi,
        'firma': firma,
        'title': title,
        'iade_sayisi': iade_sayisi,
        'ort_desi': ort_desi,
        'ort_tutar': ort_tutar,
        'top_tutar': top_tutar,
        'top_stok': top_stok,
        'top_adet': top_adet,
    }
    

    return render(request, 'etikom/iadelistesi.html', context)


def iadeexceliyuklemeyap(request):
    title = 'Excel Yükle'
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if request.method == "POST":
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                siparis_no = row['Sipariş No']
                stk_no = row['Stok Kodu']
                if isinstance(siparis_no, int):
                    siparis = str(siparis_no)
                else:
                    siparis = siparis_no
                iade = Iade(
                    Siparisno = siparis,
                    Stokkodu = row['Stok Kodu'],
                    Adet = row['Adet'],
                    Desi = row['Desi'],
                    Iadetutari = row['İade Kargo Tutarı'],
                    Firmaadi = request.user,
                    Tur = 'İ'
                )
                iade.save6()

                fyt = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=siparis, Stokkodu=stk_no).values('Satisfiyati').first()
                fiyat = fyt['Satisfiyati'] if fyt else None
                stok = Stok(
                    Firmaadi = request.user,
                    Afaturano = siparis,
                    Stokkodu = row['Stok Kodu'],
                    Adet = row['Adet'],
                    Alisfiyati = fiyat,
                    Tur = 'İ'
                )
                stok.save1()

                pzr_yeri = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=siparis, Stokkodu=stk_no).values('Pazaryeri').first()
                pazaryeri = pzr_yeri['Pazaryeri']
                trh = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=siparis, Stokkodu=stk_no).values('Tarih').first()
                tarih = trh['Tarih']
                kmsyn = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=siparis, Stokkodu=stk_no).values('Komisyon').first()
                komisyon = kmsyn['Komisyon']
                sip = Siparis(
                    Pazaryeri = pazaryeri,
                    Tarih = tarih,
                    Siparisno = row['Sipariş No'],
                    Stokkodu = row['Stok Kodu'],
                    Adet = row['Adet'] * -1,
                    Satisfiyati = fiyat,
                    Komisyon = komisyon,
                    Firmaadi = request.user,
                    Tur = 'İ'
                )
                sip.save3()

                kargo = Kargo(
                    Desi = row['Desi'],
                    Kargotutari = row['İade Kargo Tutarı'],
                    Hizmetbedeli = 0,
                    Islembedeli = 0,
                    Siparisno = siparis,
                    Stokkodu = row['Stok Kodu'],
                    Firmaadi = request.user,
                    Tur = 'İ'
                )
                kargo.save5()
            
            return redirect('iadelistesiurl')

    return render(request, 'etikom/iadeexceliyukle.html', {'firma_adi': firma_adi, 'title': title})

def iadeduzeltme(request, firma, pk):
    
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if firma_adi != firma:
        return redirect('cikisurl')

    fa = firma
    pk = pk

    kontroliade = get_object_or_404(Iade, pk=pk)
    form = IadeFormu(instance=kontroliade)


    if request.method == "POST":
        form = IadeFormu(request.POST, instance=kontroliade)
        if 'iadesil' in request.POST:
            kargo_entry = Kargo.objects.filter(Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='İ').first()
            if kargo_entry:
                kargo_entry.delete()


            siparis_entry = Siparis.objects.filter(Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='İ').first()
            if siparis_entry:
                siparis_entry.delete()


            stok_entry = Stok.objects.filter(Afaturano=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='İ').first()
            if stok_entry:
                stok_entry.delete()

            kontroliade.delete()
            return redirect('iadelistesiurl')
        elif 'iadeekle' in request.POST:
            kargo_entry = Kargo.objects.filter(Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='İ').first()
            if kargo_entry:
                kargo_entry.delete()


            siparis_entry = Siparis.objects.filter(Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='İ').first()
            if siparis_entry:
                siparis_entry.delete()


            stok_entry = Stok.objects.filter(Afaturano=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='İ').first()
            if stok_entry:
                stok_entry.delete()

            kontroliade.delete()
            if form.is_valid():
                post = form.save(commit=False)
                post.Firmaadi = request.user
                post.Tur = 'İ'
                post.save6()


                desi = form.cleaned_data['Desi']
                iade_ttr = form.cleaned_data['Iadetutari']
                kargo = Kargo(Firmaadi=request.user, Tur='İ', Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Desi=desi, Hizmetbedeli=0, Kargotutari=iade_ttr)
                kargo.save5()


                Firmaadi = request.user
                adet = form.cleaned_data['Adet']
                fyt = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu).values('Satisfiyati').first()
                fiyat = fyt['Satisfiyati']
                stok = Stok(Firmaadi=request.user, Afaturano=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Adet=adet, Alisfiyati=fiyat, Tur='İ')
                stok.save1()


                pzr_yeri = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='S').values('Pazaryeri').first()
                pazaryeri = pzr_yeri['Pazaryeri']
                adt = form.cleaned_data['Adet'] * -1
                trh = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='S').values('Tarih').first()
                tarih = trh['Tarih']
                kmsyn = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Tur='S').values('Komisyon').first()
                komisyon = kmsyn['Komisyon']
                siparis = Siparis(Firmaadi=request.user, Pazaryeri=pazaryeri, Tarih=tarih, Siparisno=kontroliade.Siparisno, Stokkodu=kontroliade.Stokkodu, Adet=adt, Komisyon=komisyon, Satisfiyati=fiyat, Tur='İ')
                siparis.save4()


                return redirect('iadelistesiurl')
    else:
        form = IadeFormu(instance=kontroliade)

    title = 'İade Detayı'
    
    return render(request, 'etikom/iadeduzeltme.html', {'form': form, 'firma_adi': firma_adi, 'title': title})

def gidereklemeyap(request):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Gider'
    
    if request.method == "POST":
        if 'giderekle' in request.POST:
            form = GiderFormu(request.POST)
            if form.is_valid():
                post = form.save(commit=False)
                post.Firmaadi = request.user
                post.save()
                return redirect('giderlistesiurl')
                    
    else:
        form = GiderFormu()

    return render(request, 'etikom/giderekle.html', {'title': title, 'firma_adi': firma_adi, 'form': form})

def giderlistesiyap(request, sort=None):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    gider = Gider.objects.filter(Firmaadi=firma_adi_id).count()
    if gider is None:
        gider_sayisi = 0
    else:
        gider_sayisi = gider

    basliklar = Gider.objects.filter(Firmaadi=firma_adi_id).values('Baslik').distinct().count()                  # baslik sayisi
    tutarlar = Gider.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Tutar"))["Tutar__sum"]                 # toplam tutar


    if gider_sayisi == 0:
        baslik_sayisi = 0
        top_tutar = 0
        ort_tutar = 0
    else:
        baslik_sayisi = basliklar
        top_tutar = tutarlar
        ort_tutar = top_tutar/baslik_sayisi
    

    if sort == 'az-baslik':
        gider = Gider.objects.filter(Firmaadi=firma_adi_id).order_by('Baslik').values()
    elif sort == 'za-baslik':
        gider = Gider.objects.filter(Firmaadi=firma_adi_id).order_by('-Baslik').values()
    elif sort == 'az-tarih':
        gider = Gider.objects.filter(Firmaadi=firma_adi_id).order_by('Tarih').values()
    elif sort == 'za-tarih':
        gider = Gider.objects.filter(Firmaadi=firma_adi_id).order_by('-Tarih').values()
    elif sort == 'az-tutar':
        gider = Gider.objects.filter(Firmaadi=firma_adi_id).order_by('Tutar').values()
    elif sort == 'za-tutar':
        gider = Gider.objects.filter(Firmaadi=firma_adi_id).order_by('-Tutar').values()
    else:
        gider = Gider.objects.filter(Firmaadi=firma_adi_id)

    firma = request.user.username
    title = 'Gider Listesi'

    context = {
        'gider': gider,
        'firma_adi': firma_adi,
        'firma': firma,
        'title': title,
        'gider_sayisi': gider_sayisi,
        'baslik_sayisi': baslik_sayisi,
        'top_tutar': top_tutar,
        'ort_tutar': ort_tutar,
    }
    

    return render(request, 'etikom/giderlistesi.html', context)

def iadeexceliindir(request):
    firma_adi_id = request.user.id

    # Kargo modelinden tüm verileri al
    iadeler = Iade.objects.filter(Firmaadi=firma_adi_id)

    # Kargo verilerini bir DataFrame'e dönüştür
    data = {
        'Sipariş No': [iade.Siparisno for iade in iadeler],
        'Stok Kodu': [iade.Stokkodu for iade in iadeler],
        'Adet': [iade.Adet for iade in iadeler],
        'Desi': [iade.Desi for iade in iadeler],
        'İade Tutarı': [str(iade.Iadetutari).replace('.', ',') for iade in iadeler],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)  # Buffer'ın başına git

    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_iade_kargo_listesi.xlsx"'

    return response

def giderduzeltme(request, firma, pk):
    
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    if firma_adi != firma:
        return redirect('cikisurl')

    fa = firma
    pk = pk

    kontrolgider = get_object_or_404(Gider, pk=pk)
    form = GiderFormu(instance=kontrolgider)


    if request.method == "POST":
        form = GiderFormu(request.POST, instance=kontrolgider)
        if 'gidersil' in request.POST:
            kontrolgider.delete()
            return redirect('giderlistesiurl')
        elif 'giderekle' in request.POST:
            if form.is_valid():
                post = form.save(commit=False)
                post.Firmaadi = request.user
                post.save()
                return redirect('giderlistesiurl')
    else:
        form = GiderFormu(instance=kontrolgider)

    title = 'Gider Detayı'
    
    return render(request, 'etikom/giderduzeltme.html', {'form': form, 'firma_adi': firma_adi, 'title': title})

def giderexceliindir(request):
    firma_adi_id = request.user.id

    # Kargo modelinden tüm verileri al
    giderler = Gider.objects.filter(Firmaadi=firma_adi_id)

    # Kargo verilerini bir DataFrame'e dönüştür
    data = {
        'Başlık': [gider.Baslik for gider in giderler],
        'Tarih': [gider.Tarih.strftime('%d.%m.%Y') for gider in giderler],
        'Tutar': [str(gider.Tutar).replace('.', ',') for gider in giderler],
    }
    df = pd.DataFrame(data)

    # DataFrame'i Excel dosyasına dönüştür
    excel_buffer = BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)  # Buffer'ın başına git

    # HTTP yanıtı olarak Excel dosyasını döndür
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{request.user.username}_etikom_gider_listesi.xlsx"'

    return response

def giderexceliyuklemeyap(request):
    title = 'Excel Yükle'
    firma_adi = request.user.username

    if request.method == "POST":
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file)
            for index, row in df.iterrows():
                gider = Gider(
                    Baslik = row['Başlık'],
                    Tarih = row['Tarih'],
                    Tutar = row['Tutar'],
                    Firmaadi = request.user,
                )
                gider.save()
            
            return redirect('giderlistesiurl')

    return render(request, 'etikom/giderexceliyukle.html', {'firma_adi': firma_adi, 'title': title})

def giderbaslikdetayiyap(request, sort):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Baslık Detayı'
    baslik = Gider.objects.filter(Firmaadi=firma_adi_id, Baslik=sort)

    baslik_sayisi = baslik.count()                                                                               # baslik sayisi
    top_tutar = baslik.aggregate(Sum("Tutar"))["Tutar__sum"]                                                      # toplam tutar

    context = {
        'firma_adi': firma_adi,
        'title': title,
        'sort': sort,
        'baslik': baslik,
        'baslik_sayisi': baslik_sayisi,
        'top_tutar': top_tutar,
    }

    return render(request, 'etikom/giderbaslikdetayi.html', context)

def siparisdetayiyap(request, sort):
    firma_adi = request.user.username
    firma_adi_id = request.user.id
    title = 'Sipariş Detayı'

    siparisler = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sort)               # siparis modelinde ki ayni nolu siparisler

    tstc = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sort).values('Stokkodu').order_by('Stokkodu').distinct().count()
    tsta = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sort).aggregate(Sum("Adet"))["Adet__sum"]
    
    tsta = abs(tsta)

    sftt = Siparis.objects.filter(Firmaadi=firma_adi_id, Siparisno=sort).aggregate(Sum("Toplam"))["Toplam__sum"]

    sftt = abs(sftt)

    oafi = sftt / tsta

    context = {
        'firma_adi': firma_adi,
        'title': title,
        'siparisler': siparisler,
        'tstc': tstc,
        'tsta': tsta,
        'sftt': sftt,
        'oafi': oafi,
        'sort': sort,
    }

    return render(request, 'etikom/siparisdetayi.html', context)

def faturakontrolyap(request, sort=None):
    firma_adi = request.user.username
    firma_adi_id = request.user.id

    tpys = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Pazaryeri').order_by('Pazaryeri').distinct().count()
    tsps = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Siparisno').order_by('Siparisno').distinct().count()   # toplam siparis sayisi
    tsts = Siparis.objects.filter(Firmaadi=firma_adi_id).values('Stokkodu').order_by('Stokkodu').distinct().count()
    tstc = Siparis.objects.filter(Firmaadi=firma_adi_id, Adet__gt=0).aggregate(Sum('Adet'))["Adet__sum"]                # tum siparislerdeki toplam urun adedi

    if tstc is None:
        tstc = 0

    tsius = Siparis.objects.filter(Firmaadi=firma_adi_id, Adet__lte=0).aggregate(Sum('Adet'))["Adet__sum"]                # tum siparislerdeki iade urun sayısı

    if tsius is None:
        ius = 0
    else:
        ius = abs(tsius)

    if tstc == 0:                   
        orsp = 0
    else:
        orsp = tstc / tsps          # siparis basina dusen urun sayisi

    tstt = Siparis.objects.filter(Firmaadi=firma_adi_id).aggregate(Sum("Toplam"))["Toplam__sum"]

    if tstt is None:
        tstt = 0
        ostt = 0
    else:
        ostt = tstt / tsps

    stsys = Siparis.objects.filter(Firmaadi=firma_adi_id).count()

    if sort == 'az-tur':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Tur').values()
    elif sort == 'za-tur':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Tur').values()
    elif sort == 'az-pazaryeri':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Pazaryeri').values()
    elif sort == 'za-pazaryeri':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Pazaryeri').values()
    elif sort == 'az-tarih':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Tarih').values()
    elif sort == 'za-tarih':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Tarih').values()
    elif sort == 'az-siparis-no':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Siparisno').values()
    elif sort == 'za-siparis-no':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Siparisno').values()
    elif sort == 'az-stok-kodu':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Stokkodu').values()
    elif sort == 'za-stok-kodu':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Stokkodu').values()
    elif sort == 'az-adet':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Adet').values()
    elif sort == 'za-adet':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Adet').values()
    elif sort == 'az-satis-fiyati':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Satisfiyati').values()
    elif sort == 'za-satis-fiyati':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Satisfiyati').values()
    elif sort == 'az-toplam':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Toplam').values()
    elif sort == 'za-toplam':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Toplam').values()
    elif sort == 'az-komisyon-orani':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Komisyon').values()
    elif sort == 'za-komisyon-orani':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Komisyon').values()
    elif sort == 'az-komisyon-tutari':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Komisyontutari').values()
    elif sort == 'za-komisyon-tutari':
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('-Komisyontutari').values()
    else:
        siparis = Siparis.objects.filter(Firmaadi=firma_adi_id).order_by('Tarih')[:10]

    firma = request.user.username
    title = 'Fatura Kontrol'

    context = {
        'siparis': siparis,
        'firma_adi': firma_adi,
        'title': title,
        'tpys': tpys,
        'tsps': tsps,
        'tsts': tsts,
        'tstc': tstc,
        'ius': ius,
        'orsp': orsp,
        'tstt': tstt,
        'ostt': ostt,
        'stsys': stsys,
        'firma': firma,
    }

    return render(request, 'etikom/faturakontrol.html', context)